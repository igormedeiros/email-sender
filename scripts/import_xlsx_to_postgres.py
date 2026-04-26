#!/usr/bin/env python3
"""
Script de importação de contatos a partir de arquivos Excel (.xlsx).
- Valida emails
- Insere client_name e email no banco
- Adiciona tag opcional aos contatos importados
- Verifica duplicatas

Dependência: openpyxl (já instalada no venv)

Uso:
    python scripts/import_xlsx_to_postgres.py to_import/Mailing\ Pernambuco_filtrado.xlsx --tag "Mailing Pernambuco"
"""

import sys
import re
import logging
import argparse
from pathlib import Path
import os
import psycopg
from psycopg.rows import dict_row
import openpyxl

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(levelname)s] %(message)s'
)
log = logging.getLogger(__name__)

# Regex para validar email
EMAIL_REGEX = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'


class ExcelImporter:
    """Importador de Excel para o Banco Postgres."""
    
    def __init__(self, db_host, db_port, db_user, db_password, db_name):
        self.conn_params = {
            'host': db_host,
            'port': db_port,
            'user': db_user,
            'password': db_password,
            'dbname': db_name,
        }
        self.conn = None
        self.stats = {
            'total': 0,
            'valid': 0,
            'inserted': 0,
            'duplicates': 0,
            'invalid': 0,
            'errors': []
        }

    def connect(self):
        try:
            self.conn = psycopg.connect(**self.conn_params)
            log.info(f"✅ Conectado ao banco: {self.conn_params['dbname']}")
        except Exception as e:
            log.error(f"❌ Erro ao conectar: {e}")
            sys.exit(1)

    def close(self):
        if self.conn:
            self.conn.close()

    def validate_email(self, email):
        if not email:
            return False
        return bool(re.match(EMAIL_REGEX, str(email).strip().lower()))

    def get_or_create_tag(self, tag_name):
        if not tag_name:
            return None
        
        try:
            with self.conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO tbl_tags (tag_name) VALUES (%s) ON CONFLICT (tag_name) DO NOTHING RETURNING id",
                    (tag_name,)
                )
                res = cur.fetchone()
                if res:
                    return res[0]
                
                cur.execute("SELECT id FROM tbl_tags WHERE tag_name = %s", (tag_name,))
                res = cur.fetchone()
                return res[0] if res else None
        except Exception as e:
            log.error(f"Erro ao obter/criar tag {tag_name}: {e}")
            return None

    def add_tag_to_contact(self, contact_id, tag_id):
        if not tag_id or not contact_id:
            return
        try:
            with self.conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO tbl_contact_tags (contact_id, tag_id) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                    (contact_id, tag_id)
                )
        except Exception as e:
            log.error(f"Erro ao adicionar tag ao contato {contact_id}: {e}")

    def import_from_xlsx(self, file_path, tag_name=None):
        path = Path(file_path)
        if not path.exists():
            log.error(f"Arquivo não encontrado: {file_path}")
            return

        try:
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook.active
            
            # Mapear colunas (Headers)
            headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            
            # Tentar encontrar colunas de email e nome
            email_idx = -1
            name_idx = -1
            
            # Possíveis nomes para colunas
            email_aliases = ['e-mail', 'email', 'contato', 'correio']
            name_aliases = ['nome', 'client_name', 'cliente', 'name']
            
            for idx, h in enumerate(headers):
                if h in email_aliases:
                    email_idx = idx
                elif h in name_aliases:
                    name_idx = idx
            
            if email_idx == -1:
                log.error("Coluna de email não encontrada. Esperado: 'E-mail' ou 'Email'")
                return

            tag_id = self.get_or_create_tag(tag_name) if tag_name else None
            
            log.info(f"🚀 Iniciando importação de {file_path}...")
            
            count = 0
            for row in sheet.iter_rows(min_row=2):
                self.stats['total'] += 1
                email = row[email_idx].value
                name = row[name_idx].value if name_idx != -1 else None
                
                if not email:
                    continue
                
                email = str(email).strip().lower()
                name = str(name).strip() if name else None
                
                if not self.validate_email(email):
                    self.stats['invalid'] += 1
                    continue
                
                self.stats['valid'] += 1
                
                try:
                    with self.conn.cursor() as cur:
                        # Inserir ou atualizar nome (opcional)
                        cur.execute(
                            """
                            INSERT INTO tbl_contacts (email, client_name, unsubscribed, is_buyer)
                            VALUES (%s, %s, FALSE, FALSE)
                            ON CONFLICT (email) DO UPDATE 
                            SET client_name = COALESCE(EXCLUDED.client_name, tbl_contacts.client_name)
                            RETURNING id, (xmin = 0) as is_new
                            """,
                            (email, name)
                        )
                        # Nota: Em Postgres, para saber se foi inserido ou atualizado via ON CONFLICT
                        # podemos usar algumas técnicas, mas aqui o importante é ter o ID
                        res = cur.fetchone()
                        if res:
                            contact_id = res[0]
                            self.stats['inserted'] += 1
                            
                            if tag_id:
                                self.add_tag_to_contact(contact_id, tag_id)
                        
                except Exception as e:
                    self.stats['errors'].append(f"Erro ao processar {email}: {e}")
                    log.error(f"Erro em {email}: {e}")
                
                count += 1
                if count % 100 == 0:
                    self.conn.commit()
                    log.info(f"Processed {count} rows...")

            self.conn.commit()
            log.info("🏁 Importação concluída!")
            self.print_summary()

        except Exception as e:
            log.error(f"Erro ao processar arquivo Excel: {e}")
        finally:
            workbook.close()

    def print_summary(self):
        log.info("\n" + "="*40)
        log.info("📊 RESUMO DA IMPORTAÇÃO")
        log.info("="*40)
        log.info(f"Total rows:      {self.stats['total']}")
        log.info(f"E-mails válidos: {self.stats['valid']}")
        log.info(f"Processados:     {self.stats['inserted']}")
        log.info(f"Inválidos:       {self.stats['invalid']}")
        log.info(f"Erros:           {len(self.stats['errors'])}")
        log.info("="*40)


def main():
    parser = argparse.ArgumentParser(description="Importador XLSX para Postgres")
    parser.add_argument("file", help="Caminho do arquivo .xlsx")
    parser.add_argument("--tag", help="Tag para associar aos contatos", default=None)
    args = parser.parse_args()

    # Carregar .env
    env_file = Path(".env")
    if not env_file.exists():
        log.error("❌ Arquivo .env não encontrado")
        sys.exit(1)
    
    env_vars = {}
    with open(env_file, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#'):
                key, value = line.split('=', 1)
                env_vars[key.strip()] = value.strip()
    
    db_host = env_vars.get('PGHOST', 'localhost')
    db_port = int(env_vars.get('PGPORT', 5432))
    db_user = env_vars.get('PGUSER', 'postgres')
    db_password = env_vars.get('PGPASSWORD', '')
    db_name = env_vars.get('PGDATABASE', 'treineinsite')

    importer = ExcelImporter(db_host, db_port, db_user, db_password, db_name)
    importer.connect()
    try:
        importer.import_from_xlsx(args.file, args.tag)
    finally:
        importer.close()


if __name__ == "__main__":
    main()
